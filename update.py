from openpyxl import load_workbook
import json
import os
import subprocess
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CUSTOMERS_FILE = os.path.join(BASE_DIR, "customers.xlsx")
CUSTOMERS_FOLDER = os.path.join(BASE_DIR, "Customers")
JSON_FILE = os.path.join(BASE_DIR, "customers.json")


def format_money(value):
    if value is None:
        return "0 تومان"

    try:
        value = int(float(value))
        return f"{value:,}".replace(",", ".") + " تومان"
    except:
        return str(value)


wb = load_workbook(CUSTOMERS_FILE)
ws = wb.active

customers = {}
for row in range(2, ws.max_row + 1):

    code = ws[f"A{row}"].value
    name = ws[f"B{row}"].value
    date = ws[f"D{row}"].value
    filename = ws[f"E{row}"].value

    if code is None or filename is None:
        continue

    customer_file = os.path.join(CUSTOMERS_FOLDER, str(filename))

    amount = 0

    if os.path.exists(customer_file):

        try:

            customer_wb = load_workbook(customer_file, data_only=True)
            customer_ws = customer_wb.active

            amount = customer_ws["H2"].value

            if amount is None:
                amount = 0

        except Exception as e:

            print("Error reading file.")
            print(e)

    else:

        print("file not found.")

    ws[f"C{row}"] = amount

    customers[str(code)] = {
        "name": str(name),
        "money": format_money(amount),
        "date": str(date)
    }
    wb.save(CUSTOMERS_FILE)

with open(JSON_FILE, "w", encoding="utf-8") as f:
    json.dump(customers, f, ensure_ascii=False, indent=4)

print("customers.json created.")

msg = "Update " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")

try:

    subprocess.run(["git", "add", "."], check=True)

    result = subprocess.run(
        ["git", "commit", "-m", msg],
        capture_output=True,
        text=True
    )

    if result.returncode != 0:
        print("not change for commit.")

    subprocess.run(["git", "pull", "--rebase", "origin", "main"], check=True)

    subprocess.run(["git", "push", "origin", "main"], check=True)

    print("")
    print("===================================")
    print("GitHub Updated Successfully.")
    print("===================================")

except Exception as e:

    print("")
    print("Github error")
    print(e)
